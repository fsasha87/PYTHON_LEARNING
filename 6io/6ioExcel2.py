# xlsx: p_i_openpyxl: wb=openpyxl.load_workbook(path)->s=wb.active->print(s.cell(1,2).value)->print(s['A2'].value)
# ->for_row_in_s.iter_rows(...):for_c_in_row:print(c.value); s['A1']='Hi'->wb.save(path)
import openpyxl
path = "../testFiles/data1.xlsx"
workbook = openpyxl.load_workbook(path)
sheet_obj = workbook.active
cell_obj = sheet_obj.cell(row=1, column=2)

print(cell_obj.value)
print(sheet_obj['C3'].value)

for row in sheet_obj.iter_rows(min_row=1, max_row=5, min_col=1, max_col=5):
    for cell in row:
        print(cell.value)


# Записать данные в конкретные ячейки
sheet_obj['D1'] = 'Hello'
sheet_obj['E1'] = 'World'
sheet_obj['D2'] = 123
sheet_obj['E2'] = 456

workbook.save('testFiles/data1.xlsx')
